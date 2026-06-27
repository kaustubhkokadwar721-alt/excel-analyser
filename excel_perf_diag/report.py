"""Four-sheet report writer with the presentation layer (spec §10–11).

Summary (health strip + Top-5) / Formula Cost (Pareto, data bars, cum %) /
Structure / Actions (ROI-ranked). Written to a SEPARATE report workbook so the
analyzed file is never mutated.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

GREEN = "C6EFCE"
AMBER = "FFEB9C"
RED = "FFC7CE"
HDR = "1F4E78"
GREY = "F2F2F2"

_hdr_font = Font(color="FFFFFF", bold=True)
_hdr_fill = PatternFill("solid", fgColor=HDR)
_bold = Font(bold=True)
_thin = Side(style="thin", color="D9D9D9")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _band(value, warn, bad, reverse=False):
    """Return a fill color. reverse=True means LOW is bad (e.g. MT efficiency)."""
    if value is None:
        return GREY
    if reverse:
        if value <= bad:
            return RED
        if value <= warn:
            return AMBER
        return GREEN
    if value >= bad:
        return RED
    if value >= warn:
        return AMBER
    return GREEN


def _hcell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _hdr_font
    c.fill = _hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border
    return c


def _flag_reason(p):
    bits = []
    if p.is_volatile:
        fn = next((f for f in p.funcs if f in
                   {"INDIRECT", "OFFSET", "NOW", "TODAY", "RAND", "RANDBETWEEN", "CELL", "INFO"}), "volatile")
        if fn == "INDIRECT":
            bits.append("INDIRECT — volatile + single-threaded + ref-resolution cost")
        else:
            bits.append(f"{fn} — volatile, recalcs on every change")
    if p.cells_touched >= 1_048_576:
        bits.append("full-column reference — scans whole column")
    if p.is_single_threaded and "INDIRECT" not in p.funcs:
        bits.append("single-threaded — blocks other cores")
    if p.func_class == "array":
        bits.append("array/SUMPRODUCT — cost scales with cells touched")
    if p.nonlinear:
        bits.append("non-linear cost vs size")
    return "; ".join(bits)


# ---------------------------------------------------------------- Summary
def _write_summary(wb, report):
    ws = wb.active
    ws.title = "Summary"
    s = report.summary
    ws["A1"] = "Excel Performance Diagnostic — Summary"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = f"File: {s.file}"
    ws["A3"] = (f"Excel {s.excel_version} · {s.machine} · {s.timestamp}"
                + ("  ·  PARTIAL (time budget hit)" if s.partial else ""))
    ws["A3"].font = Font(italic=True, color="808080")

    # Health strip
    r = 5
    _hcell(ws, r, 1, "Health strip — colour driven by measured threshold")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    strip = [
        ("Volatility %", s.volatility_pct, f"{s.volatility_pct:.0f}%",
         _band(s.volatility_pct, 40, 60)),
        ("MT efficiency", s.mt_efficiency, f"{s.mt_efficiency:.2f}×",
         _band(s.mt_efficiency, 2.5, 1.5, reverse=True)),
        ("Used-range waste", s.used_range_waste_pct, f"{s.used_range_waste_pct:.0f}%",
         _band(s.used_range_waste_pct, 50, 90)),
        ("Open time", s.open_ms, f"{s.open_ms/1000:.1f}s",
         _band(s.open_ms, 5000, 15000)),
    ]
    for i, (label, _v, disp, color) in enumerate(strip):
        c1 = ws.cell(row=r, column=1 + i, value=label)
        c1.font = _bold
        c1.alignment = Alignment(horizontal="center")
        c2 = ws.cell(row=r + 1, column=1 + i, value=disp)
        c2.fill = PatternFill("solid", fgColor=color)
        c2.alignment = Alignment(horizontal="center")
        c2.font = Font(bold=True)
        c2.border = _border
    r += 3

    # Key metrics
    _hcell(ws, r, 1, "Metric")
    _hcell(ws, r, 2, "Value")
    metrics = [
        ("Full-calc (multi-thread)", f"{s.full_calc_ms:.0f} ms"),
        ("Recalc (volatiles+deps)", f"{s.recalc_ms:.0f} ms"),
        ("Single-thread full-calc", f"{s.single_thread_ms:.0f} ms"),
        ("Volatility %", f"{s.volatility_pct:.1f}%"),
        ("Multi-thread efficiency", f"{s.mt_efficiency:.2f}×"),
        ("Max dependency depth", s.max_dependency_depth),
        ("Open / Save", f"{s.open_ms:.0f} / {s.save_ms:.0f} ms"),
        ("File size", f"{s.file_size_mb:.2f} MB"),
        ("Used-range waste (worst sheet)", f"{s.used_range_waste_pct:.0f}%"),
        ("Cell-format count (styles.xml)", s.style_count),
        ("External links", s.external_link_count),
        ("Defined names", s.defined_name_count),
        ("PQ total refresh", f"{s.pq_total_refresh_s:.2f} s"),
    ]
    for i, (k, v) in enumerate(metrics):
        ws.cell(row=r + 1 + i, column=1, value=k).border = _border
        ws.cell(row=r + 1 + i, column=2, value=v).border = _border
    r2 = r + 1 + len(metrics) + 1

    # Top 5 costs callout
    measured = sorted([p for p in report.patterns if p.measured and p.total_ms > 0],
                      key=lambda x: -x.total_ms)
    total = sum(p.total_ms for p in measured) or 1.0
    _hcell(ws, r2, 1, "Top costs")
    _hcell(ws, r2, 2, "Total ms")
    _hcell(ws, r2, 3, "Share")
    _hcell(ws, r2, 4, "Why")
    cum = 0.0
    for i, p in enumerate(measured[:5]):
        cum += p.total_ms
        ws.cell(row=r2 + 1 + i, column=1, value=f"{p.sheet}!{p.sample_cell} {p.func_class}")
        ws.cell(row=r2 + 1 + i, column=2, value=round(p.total_ms, 1))
        ws.cell(row=r2 + 1 + i, column=3, value=f"{100*p.total_ms/total:.0f}%")
        ws.cell(row=r2 + 1 + i, column=4, value=_flag_reason(p) or p.func_class)
    if measured:
        ws.cell(row=r2 + 6, column=1, value=f"→ Top 5 = {100*cum/total:.0f}% of measured calc cost")
        ws.cell(row=r2 + 6, column=1).font = _bold

    # Regression deltas
    if report.deltas and report.deltas.get("summary"):
        rr = r2 + 8
        _hcell(ws, rr, 1, f"Change vs previous run ({report.deltas.get('prev_timestamp','')})")
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
        for i, (k, d) in enumerate(report.deltas["summary"].items()):
            pct = f"{d['pct']:+.0f}%" if d.get("pct") is not None else ""
            cell = ws.cell(row=rr + 1 + i, column=1,
                           value=f"{k}: {d['prev']} → {d['now']}  ({d['delta']:+.1f}, {pct})")
            if d["delta"] < 0:
                cell.fill = PatternFill("solid", fgColor=GREEN)
            elif d["delta"] > 0:
                cell.fill = PatternFill("solid", fgColor=RED)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------- Formula Cost
def _write_formula_cost(wb, report):
    ws = wb.create_sheet("Formula Cost")
    cols = ["Sheet", "Pattern (R1C1)", "Class", "Occurrences", "Cells/occ",
            "µs/occ", "Stdev µs", "Total ms", "Cumulative %", "% of sheet",
            "Volatile?", "Single-thread?", "Flag + reason"]
    for i, c in enumerate(cols, 1):
        _hcell(ws, 1, i, c)

    measured = [p for p in report.patterns if p.measured]
    unmeasured = [p for p in report.patterns if not p.measured]
    measured.sort(key=lambda x: -x.total_ms)
    grand = sum(p.total_ms for p in measured) or 1.0
    sheet_tot = {}
    for p in measured:
        sheet_tot[p.sheet] = sheet_tot.get(p.sheet, 0.0) + p.total_ms

    row = 2
    cum = 0.0
    for p in measured:
        cum += p.total_ms
        vals = [p.sheet, p.r1c1[:120], p.func_class, p.occurrences, p.cells_touched,
                round(p.us_per_occ, 2), round(p.stdev_us, 2), round(p.total_ms, 2),
                round(100 * cum / grand, 1),
                round(100 * p.total_ms / sheet_tot.get(p.sheet, 1), 1),
                "Y" if p.is_volatile else "", "Y" if p.is_single_threaded else "",
                _flag_reason(p)]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = _border
        row += 1
    # unmeasured (not in top-K sheets / budget) listed below, greyed
    for p in sorted(unmeasured, key=lambda x: -x.occurrences)[:200]:
        vals = [p.sheet, p.r1c1[:120], p.func_class, p.occurrences, p.cells_touched,
                "", "", "", "", "", "Y" if p.is_volatile else "",
                "Y" if p.is_single_threaded else "", "(not timed: static suspect)"]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.fill = PatternFill("solid", fgColor=GREY)
        row += 1

    last = row - 1
    if last >= 2:
        # data bar on Total ms (col 8)
        rule = DataBarRule(start_type="num", start_value=0, end_type="max",
                           color="638EC6", showValue=True)
        ws.conditional_formatting.add(f"H2:H{last}", rule)
        ws.auto_filter.ref = f"A1:M{last}"
    ws.freeze_panes = "A2"
    widths = [16, 40, 13, 12, 11, 9, 9, 10, 12, 10, 9, 12, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------- Structure
def _write_structure(wb, report):
    ws = wb.create_sheet("Structure")
    s = report.summary
    cols = ["Sheet", "Dimension", "Populated cells", "Used-range waste %",
            "CF rules", "CF full-column?", "Max chain depth"]
    for i, c in enumerate(cols, 1):
        _hcell(ws, 1, i, c)
    row = 2
    for st in sorted(report.structure, key=lambda x: -x.used_range_waste_pct):
        vals = [st.sheet, st.dimension, st.populated_cells, st.used_range_waste_pct,
                st.cf_rule_count, "Y" if st.cf_full_column else "", st.max_chain_depth]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = _border
            if i == 4 and isinstance(v, (int, float)):
                cell.fill = PatternFill("solid", fgColor=_band(v, 50, 90))
        row += 1
    row += 1
    extras = [
        ("Cell-format count (styles.xml)", s.style_count, s.style_count > 5000),
        ("External links", s.external_link_count, s.external_link_count > 0),
        ("Defined names", s.defined_name_count, False),
        ("PQ total refresh (s)", s.pq_total_refresh_s, s.pq_total_refresh_s > 5),
    ]
    for k, v, warn in extras:
        ws.cell(row=row, column=1, value=k).font = _bold
        c = ws.cell(row=row, column=2, value=v)
        if warn:
            c.fill = PatternFill("solid", fgColor=AMBER)
        row += 1
    # Power query detail
    if report.power_queries:
        row += 1
        _hcell(ws, row, 1, "Power Query")
        _hcell(ws, row, 2, "Refresh (s)")
        _hcell(ws, row, 3, "Note")
        row += 1
        for q in report.power_queries:
            ws.cell(row=row, column=1, value=q.name)
            ws.cell(row=row, column=2, value=q.refresh_s)
            ws.cell(row=row, column=3, value=q.note)
            row += 1
    for i, w in enumerate([20, 16, 16, 18, 10, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------- Actions
def _write_actions(wb, report):
    ws = wb.create_sheet("Actions")
    ws["A1"] = "Actions — ranked by measured ROI (each anchored to evidence)"
    ws["A1"].font = Font(bold=True, size=13)
    cols = ["#", "Anchor (evidence)", "Measured cost", "Why", "Fix",
            "Est. gain", "Effort", "Confidence", "ROI"]
    for i, c in enumerate(cols, 1):
        _hcell(ws, 2, i, c)
    row = 3
    for n, a in enumerate(report.actions, 1):
        vals = [n, a.anchor, a.measured_cost, a.why, a.fix, a.est_gain,
                a.effort, a.confidence, a.roi]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = _border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=9).fill = PatternFill(
            "solid", fgColor=_band(a.roi, 30, 80))
        row += 1
    if not report.actions:
        ws.cell(row=3, column=2, value="No high-ROI actions — file is healthy on measured axes.")
    for i, w in enumerate([4, 30, 26, 40, 46, 18, 9, 11, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def write(report, out_path):
    wb = Workbook()
    _write_summary(wb, report)
    _write_formula_cost(wb, report)
    _write_structure(wb, report)
    _write_actions(wb, report)
    wb.save(out_path)
    return out_path
