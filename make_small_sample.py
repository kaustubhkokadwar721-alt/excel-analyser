"""Small fixture for fast .xlam verification (bounded ranges, modest counts).
Run: python make_small_sample.py  ->  PerfTest_Small.xlsx
"""
import random
from openpyxl import Workbook
random.seed(1)

N = 1500          # data rows
M = 800           # formula rows per sheet

wb = Workbook()
d = wb.active; d.title = "Data"
d.append(["ID", "Region", "Qty", "Price"])
regions = ["North", "South", "East", "West"]
for i in range(1, N + 1):
    d.append([i, random.choice(regions), random.randint(1, 500),
              round(random.uniform(5, 500), 2)])
last = N + 1

# Lookup legacy vs modern (bounded ranges - no full-column blowup)
ll = wb.create_sheet("Lookup_Legacy")
ll.append(["key", "VLOOKUP", "INDEXMATCH"])
for r in range(2, M + 2):
    ll[f"A{r}"] = random.randint(1, N)
    ll[f"B{r}"] = f"=VLOOKUP(A{r},Data!$A$2:$D${last},2,FALSE)"
    ll[f"C{r}"] = f"=INDEX(Data!$C$2:$C${last},MATCH(A{r},Data!$A$2:$A${last},0))"

lm = wb.create_sheet("Lookup_Modern")
lm.append(["key", "XLOOKUP", "XMATCH"])
for r in range(2, M + 2):
    lm[f"A{r}"] = random.randint(1, N)
    lm[f"B{r}"] = f'=_xlfn.XLOOKUP(A{r},Data!$A$2:$A${last},Data!$B$2:$B${last},"na")'
    lm[f"C{r}"] = f"=INDEX(Data!$C$2:$C${last},_xlfn.XMATCH(A{r},Data!$A$2:$A${last}))"

# Aggregation legacy vs modern
al = wb.create_sheet("Agg_Legacy")
al.append(["Region", "SUMPRODUCT"])
for i, reg in enumerate(regions, start=2):
    al[f"A{i}"] = reg
    al[f"B{i}"] = f"=SUMPRODUCT((Data!$B$2:$B${last}=A{i})*Data!$C$2:$C${last})"

am = wb.create_sheet("Agg_Modern")
am.append(["Region", "SUMIFS"])
for i, reg in enumerate(regions, start=2):
    am[f"A{i}"] = reg
    am[f"B{i}"] = f"=SUMIFS(Data!$C$2:$C${last},Data!$B$2:$B${last},A{i})"

# Volatile (relative -> collapses to one pattern each)
vol = wb.create_sheet("Volatile")
vol.append(["OFFSET", "INDIRECT", "RAND"])
for r in range(2, M + 2):
    vol[f"A{r}"] = "=SUM(OFFSET(Data!$C$1,ROW()-1,0,5,1))"
    vol[f"B{r}"] = '=INDIRECT("Data!C"&ROW())'
    vol[f"C{r}"] = "=RAND()"

wb.save("PerfTest_Small.xlsx")
print("Wrote PerfTest_Small.xlsx  rows=%d formula_rows=%d" % (N, M))
