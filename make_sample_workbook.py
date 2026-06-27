"""
make_sample_workbook.py
Generates PerfTest_Sample.xlsx — a synthetic stress workbook for the Excel
Performance Diagnostic Tool. Puts LEGACY vs MODERN formulas on the SAME data
so per-pattern cost differences can be measured and checked against known
expectations.

Run:  python make_sample_workbook.py [N_ROWS]
Default N_ROWS = 10000. Bump to 50000 for a genuine stress file on i3 hardware.

Notes on new functions: openpyxl writes formula strings verbatim, so post-2007
worksheet functions need the internal prefixes Excel stores in XML, or Excel
shows #NAME?. Helpers _f()/_fa() add the correct _xlfn. / _xlfn._xlws. prefixes.
"""

import sys
import random
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import PatternFill, Font, Border, Side, NamedStyle, Alignment
from openpyxl.utils import get_column_letter

N = int(sys.argv[1]) if len(sys.argv) > 1 else 10000
random.seed(42)

# Functions needing the _xlfn. prefix (stored form) so Excel recognises them.
XLFN = {
    "XLOOKUP", "XMATCH", "SEQUENCE", "RANDARRAY", "LET", "LAMBDA",
    "MAP", "REDUCE", "SCAN", "BYROW", "BYCOL", "MAKEARRAY",
    "MAXIFS", "MINIFS", "TEXTJOIN", "SWITCH", "IFS", "CONCAT",
    "TEXTBEFORE", "TEXTAFTER", "TEXTSPLIT",
}
# Spill worksheet functions need the _xlfn._xlws. prefix.
XLWS = {"FILTER", "SORT", "SORTBY", "UNIQUE"}


def _prefix(formula: str) -> str:
    """Add stored-XML prefixes for modern functions inside a formula string."""
    import re

    def repl(m):
        name = m.group(1)
        up = name.upper()
        nxt = m.group(2)  # the '(' that follows
        if up in XLWS:
            return f"_xlfn._xlws.{up}{nxt}"
        if up in XLFN:
            return f"_xlfn.{up}{nxt}"
        return m.group(0)

    return re.sub(r"\b([A-Za-z][A-Za-z0-9_.]*)(\()", repl, formula)


def _f(ws, cell, formula):
    """Write a normal formula (with modern-function prefixing)."""
    ws[cell] = "=" + _prefix(formula)


def _fa(ws, cell, formula, ref=None):
    """Write a legacy CSE array formula."""
    ref = ref or cell
    ws[cell] = ArrayFormula(ref, "=" + _prefix(formula))


# ---------------------------------------------------------------- styles
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border


wb = Workbook()

# ================================================================ README
readme = wb.active
readme.title = "README"
readme["A1"] = "PerfTest_Sample.xlsx — diagnostic stress / legacy-vs-modern test workbook"
readme["A1"].font = Font(bold=True, size=14)
rows = [
    ("Sheet", "Tests", "Expected signal"),
    ("Data", "Shared base dataset + Excel Table", "source of truth"),
    ("Lookup_Legacy", "VLOOKUP/INDEX-MATCH exact, nested IFERROR", "expensive (exact scan)"),
    ("Lookup_Modern", "XLOOKUP / FILTER / XMATCH, same results", "cheaper (binary/native)"),
    ("Agg_Legacy", "SUMPRODUCT, CSE array SUM(IF), DSUM", "expensive (array)"),
    ("Agg_Modern", "SUMIFS/COUNTIFS/AVERAGEIFS/MAXIFS", "cheap (native multi-thread)"),
    ("Volatile", "OFFSET/INDIRECT/NOW/TODAY/RAND/CELL/INFO", "POISON - high volatility %"),
    ("FullColumn", "A:A refs vs bounded equivalents", "full-column waste"),
    ("DynamicArrays", "FILTER/SORT/UNIQUE/SEQUENCE spill", "dynamic-array class"),
    ("Lambda", "LAMBDA + MAP/REDUCE/SCAN, named LAMBDA", "lambda class (not UDF-block)"),
    ("Tables_Structured", "[@col] structured refs vs A2*B2", "structured-ref behaviour"),
    ("DepChain", "deep chain each cell refs previous", "max chain depth driver"),
    ("CondFormat", "full-column CF + heavy rules", "dynamic CF cost"),
    ("Formats_Bloat", "many distinct cell formats", "styles.xml stress"),
    ("GhostCells", "formatting far down/right", "used-range bloat"),
]
for r, row in enumerate(rows, start=3):
    for c, val in enumerate(row, start=1):
        readme.cell(row=r, column=c, value=val)
style_header(readme, 3, row=3)
readme["A20"] = f"Generated with N_ROWS = {N}.  Legacy/Modern pairs share identical inputs."
for col, w in {"A": 20, "B": 46, "C": 32}.items():
    readme.column_dimensions[col].width = w

# ================================================================ Data
data = wb.create_sheet("Data")
headers = ["ID", "Date", "Region", "Category", "Product", "Qty", "Price", "Cost"]
data.append(headers)
style_header(data, len(headers))
regions = ["North", "South", "East", "West", "Central"]
cats = ["A", "B", "C", "D", "E"]
start = date(2023, 1, 1)
for i in range(1, N + 1):
    data.append([
        i,
        start + timedelta(days=random.randint(0, 900)),
        random.choice(regions),
        random.choice(cats),
        f"P{random.randint(1000, 9999)}",
        random.randint(1, 500),
        round(random.uniform(5, 500), 2),
        round(random.uniform(2, 400), 2),
    ])
last = N + 1
# Excel Table over the data (structured refs available as Data[col]).
tbl = Table(displayName="tblData", ref=f"A1:H{last}")
tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
)
data.add_table(tbl)
for col, w in zip("ABCDEFGH", (8, 12, 10, 10, 10, 8, 10, 10)):
    data.column_dimensions[col].width = w

# Convenience: how many formula rows the test sheets generate (cap for sanity).
M = min(N, 8000)

# ================================================================ Lookup_Legacy
ll = wb.create_sheet("Lookup_Legacy")
ll.append(["RowKey", "VLOOKUP_Region", "INDEXMATCH_Price", "Nested_IFERROR"])
style_header(ll, 4)
for r in range(2, M + 2):
    key = r  # lookup ID == row
    _f(ll, f"A{r}", f"RANDBETWEEN(1,{N})")  # volatile-ish driver to vary key
    _f(ll, f"B{r}", f"VLOOKUP(A{r},Data!$A:$H,3,FALSE)")
    _f(ll, f"C{r}", f"INDEX(Data!$G:$G,MATCH(A{r},Data!$A:$A,0))")
    _f(ll, f"D{r}", f"IFERROR(VLOOKUP(A{r},Data!$A:$H,5,FALSE),\"na\")")

# ================================================================ Lookup_Modern
lm = wb.create_sheet("Lookup_Modern")
lm.append(["RowKey", "XLOOKUP_Region", "XMATCH_Price", "FILTER_first"])
style_header(lm, 4)
for r in range(2, M + 2):
    _f(lm, f"A{r}", f"RANDBETWEEN(1,{N})")
    _f(lm, f"B{r}", f"XLOOKUP(A{r},Data!$A:$A,Data!$C:$C,\"na\")")
    _f(lm, f"C{r}", f"INDEX(Data!$G:$G,XMATCH(A{r},Data!$A:$A))")
    _f(lm, f"D{r}", f"XLOOKUP(A{r},Data!$A:$A,Data!$E:$E,\"na\")")

# ================================================================ Agg_Legacy
al = wb.create_sheet("Agg_Legacy")
al.append(["Region", "SUMPRODUCT_Qty", "ArrayCSE_Sum", "DSUM_Qty"])
style_header(al, 4)
# Criteria block for DSUM
al["F1"] = "Region"
for i, reg in enumerate(regions, start=2):
    al[f"A{i}"] = reg
    _f(al, f"B{i}", f"SUMPRODUCT((Data!$C$2:$C${last}=A{i})*Data!$F$2:$F${last})")
    _fa(al, f"C{i}", f"SUM(IF(Data!$C$2:$C${last}=A{i},Data!$F$2:$F${last}))")
    al[f"F{i}"] = reg
    _f(al, f"D{i}", f"DSUM(Data!$A$1:$H${last},6,$F$1:$F{i})")

# ================================================================ Agg_Modern
am = wb.create_sheet("Agg_Modern")
am.append(["Region", "SUMIFS_Qty", "COUNTIFS", "AVERAGEIFS", "MAXIFS"])
style_header(am, 5)
for i, reg in enumerate(regions, start=2):
    am[f"A{i}"] = reg
    _f(am, f"B{i}", f"SUMIFS(Data!$F$2:$F${last},Data!$C$2:$C${last},A{i})")
    _f(am, f"C{i}", f"COUNTIFS(Data!$C$2:$C${last},A{i})")
    _f(am, f"D{i}", f"AVERAGEIFS(Data!$F$2:$F${last},Data!$C$2:$C${last},A{i})")
    _f(am, f"E{i}", f"MAXIFS(Data!$F$2:$F${last},Data!$C$2:$C${last},A{i})")

# ================================================================ Volatile (poison)
vol = wb.create_sheet("Volatile")
vol.append(["OFFSET_sum", "INDIRECT_ref", "NOW", "TODAY", "RAND", "CELL", "INFO"])
style_header(vol, 7)
VM = min(M, 4000)
for r in range(2, VM + 2):
    _f(vol, f"A{r}", "SUM(OFFSET(Data!$F$1,ROW()-1,0,10,1))")   # relative -> one pattern
    _f(vol, f"B{r}", "INDIRECT(\"Data!F\"&ROW())")              # relative -> one pattern
    _f(vol, f"C{r}", "NOW()")
    _f(vol, f"D{r}", "TODAY()")
    _f(vol, f"E{r}", "RAND()")
    _f(vol, f"F{r}", f"CELL(\"row\",Data!A{r})")
    _f(vol, f"G{r}", "INFO(\"numfile\")")

# ================================================================ FullColumn
fc = wb.create_sheet("FullColumn")
fc.append(["FullCol_A:A", "Bounded_equiv"])
style_header(fc, 2)
for r in range(2, min(M, 2000) + 2):
    _f(fc, f"A{r}", f"SUMIF(Data!C:C,\"North\",Data!F:F)")            # full column
    _f(fc, f"B{r}", f"SUMIF(Data!$C$2:$C${last},\"North\",Data!$F$2:$F${last})")  # bounded

# ================================================================ DynamicArrays
da = wb.create_sheet("DynamicArrays")
da.append(["What", "Spill (one anchor each)"])
style_header(da, 2)
da["A2"] = "FILTER North"
_f(da, "B2", f"FILTER(Data!$A$2:$H${last},Data!$C$2:$C${last}=\"North\")")
da["A4"] = "SORT by Qty desc"
_f(da, "B4", f"SORT(Data!$A$2:$H${last},6,-1)")
da["A6"] = "UNIQUE products"
_f(da, "B6", f"UNIQUE(Data!$E$2:$E${last})")
da["A8"] = "SEQUENCE"
_f(da, "B8", "SEQUENCE(20,1)")

# ================================================================ Lambda
lam = wb.create_sheet("Lambda")
lam.append(["What", "Result"])
style_header(lam, 2)
lam["A2"] = "Inline LAMBDA (x->x*x) via MAP over Qty"
_f(lam, "B2", f"MAP(Data!$F$2:$F$101,LAMBDA(x,x*x))")
lam["A4"] = "REDUCE sum of squares"
_f(lam, "B4", "REDUCE(0,SEQUENCE(100),LAMBDA(a,b,a+b*b))")
lam["A6"] = "SCAN running total"
_f(lam, "B6", "SCAN(0,SEQUENCE(20),LAMBDA(a,b,a+b))")
# A named LAMBDA in the defined-name table (the analogue of the user's add-in).
wb.defined_names.add  # (kept simple; named LAMBDA added below via API)
try:
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names.add(DefinedName("SQUARE", attr_text="_xlfn.LAMBDA(_xlpm.x,_xlpm.x*_xlpm.x)"))
    lam["A8"] = "Named LAMBDA SQUARE(12)"
    lam["B8"] = "=SQUARE(12)"
except Exception:
    pass

# ================================================================ Tables_Structured
ts = wb.create_sheet("Tables_Structured")
ts.append(["Structured [@]", "PlainRef", "Margin_structured"])
style_header(ts, 3)
# These mirror Data row-by-row using structured vs plain refs.
TM = min(M, 5000)
for r in range(2, TM + 2):
    _f(ts, f"A{r}", f"Data[Price]")  # spilled column ref (structured)
    break  # one structured spill anchor is enough to exercise the path
for r in range(2, TM + 2):
    _f(ts, f"B{r}", f"Data!F{r}*Data!H{r}")  # plain
    _f(ts, f"C{r}", f"(Data!G{r}-Data!H{r})*Data!F{r}")

# ================================================================ DepChain
dc = wb.create_sheet("DepChain")
dc.append(["Chain"])
style_header(dc, 1)
dc["A2"] = 1
DM = min(M, 3000)
for r in range(3, DM + 2):
    _f(dc, f"A{r}", f"A{r-1}+SIN(A{r-1})")  # each depends on the previous: deep serial chain

# ================================================================ CondFormat
cf = wb.create_sheet("CondFormat")
cf.append(["Val"])
style_header(cf, 1)
CM = min(M, 5000)
for r in range(2, CM + 2):
    cf[f"A{r}"] = random.randint(0, 1000)
# Full-column color scale + a heavy CellIs rule over the whole column.
cf.conditional_formatting.add(
    "A1:A1048576",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"),
)
cf.conditional_formatting.add(
    "A1:A1048576",
    CellIsRule(operator="greaterThan", formula=["500"],
               fill=PatternFill("solid", fgColor="FFC7CE")),
)

# ================================================================ Formats_Bloat
fb = wb.create_sheet("Formats_Bloat")
fb.append(["ManyFormats"])
style_header(fb, 1)
# Force many distinct cell-format combinations to stress styles.xml.
fmts = ["0.00", "0.000", "#,##0", "0.0%", "$#,##0.00", "dd-mmm-yy",
        "0.00E+00", "[Red]0.00", "0.00;[Blue]-0.00", "# ?/?"]
colors = ["FF0000", "00AA00", "0000FF", "AA00AA", "AAAA00",
          "00AAAA", "884400", "008888", "880088", "444444"]
for r in range(2, 802):
    cell = fb[f"A{r}"]
    cell.value = random.random() * 1000
    cell.number_format = fmts[r % len(fmts)]
    cell.font = Font(color=colors[r % len(colors)],
                     bold=(r % 2 == 0), italic=(r % 3 == 0),
                     size=10 + (r % 6))
    cell.fill = PatternFill("solid", fgColor=colors[(r + 3) % len(colors)])
    cell.border = border

# ================================================================ GhostCells
gc = wb.create_sheet("GhostCells")
gc["A1"] = "Real data ends at row 10; stray formatting inflates used range."
for r in range(2, 11):
    gc[f"A{r}"] = r
# Stray formatting far away → ghost used-range bloat.
ghost = gc["AZ50000"]
ghost.fill = PatternFill("solid", fgColor="FFFF00")
gc["BA60000"] = " "  # stray value pushing dimension further

# ---------------------------------------------------------------- save
out = "PerfTest_Sample.xlsx"
wb.save(out)
print(f"Wrote {out}  (N_ROWS={N}, formula rows M={M})")
print("Sheets:", ", ".join(s.title for s in wb.worksheets))
